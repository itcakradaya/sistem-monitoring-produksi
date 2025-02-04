from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from produksi_monitoring.models import ItemDescription  # Pastikan model sudah ada

class Command(BaseCommand):
    help = 'Mengimpor deskripsi item dari file Excel ke dalam database'

    def handle(self, *args, **kwargs):
        file_path = "/home/cakra/sistem_monitoring_produksi/backend/master_item.xlsx"# Sesuaikan dengan lokasi file

        try:
            # Membaca file Excel menggunakan openpyxl
            workbook = load_workbook(filename=file_path, data_only=True)
            sheet = workbook.active  # Menggunakan sheet pertama
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error saat membaca file Excel: {e}"))
            return

        # Ambil header (baris pertama)
        headers = [cell.value for cell in sheet[1]]

        # Pastikan kolom 'Item Description' ada
        if 'Item Description' not in headers:
            self.stdout.write(self.style.ERROR("Kolom 'Item Description' tidak ditemukan di file Excel"))
            return

        # Cari indeks kolom 'Item Description'
        desc_index = headers.index('Item Description')

        # Mengumpulkan semua deskripsi item tanpa duplikasi
        item_descriptions = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Mulai dari baris ke-2 (data)
            if row[desc_index]:  # Pastikan tidak ada data kosong
                item_descriptions.add(row[desc_index])

        # Cek apakah data sudah ada di database
        existing_descriptions = set(ItemDescription.objects.values_list('description', flat=True))
        new_descriptions = [ItemDescription(description=desc) for desc in item_descriptions if desc not in existing_descriptions]

        if new_descriptions:
            ItemDescription.objects.bulk_create(new_descriptions)
            self.stdout.write(self.style.SUCCESS(f'{len(new_descriptions)} data deskripsi item berhasil diimpor!'))
        else:
            self.stdout.write(self.style.WARNING("Tidak ada data baru yang perlu diimpor."))
